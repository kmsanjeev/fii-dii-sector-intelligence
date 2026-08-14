"""Governed empirical case intake on the shared VEDA research database."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import sqlite3
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from engines.common import config as cfg

from .cases import CaseRecord, CaseRegistry, assess_quality, normalize_case
from .persistence import EVENT_TYPES

IMPORT_STATUSES = (
    "UPLOADED", "PARSED", "MAPPED", "VALIDATED", "READY_FOR_REVIEW",
    "APPROVED_FOR_INGEST", "INGESTED", "PARTIALLY_INGESTED", "REJECTED", "FAILED",
)
ELIGIBILITY_STATES = (
    "ELIGIBLE", "ELIGIBLE_WITH_CONDITIONS", "RESEARCH_ONLY", "UNVERIFIED",
    "LEAKAGE_INVALID", "DUPLICATE", "INCOMPLETE", "REJECTED",
)
CASE_CLASSES = (
    "HISTORICAL_VERIFIED", "HISTORICAL_DOCUMENTED", "HISTORICAL_USER_REPORTED",
    "WORKED_ASTROLOGY_CASE", "PRACTITIONER_CASE", "PROSPECTIVE_VERIFIED",
    "PROSPECTIVE_PENDING",
)
QUALITY_STATES = ("HIGH", "MODERATE", "LOW", "UNVERIFIED")
VERIFICATION_STATES = (
    "SYSTEM_VERIFIED", "MULTI_SOURCE_VERIFIED", "DOCUMENT_VERIFIED",
    "USER_REPORTED", "WEAKLY_VERIFIED", "UNVERIFIED", "REFERENCE_NOT_VERIFIED",
)
TEMPLATE_VERSION = "1.0"
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

TEMPLATE_FIELDS = (
    "case_external_id", "case_class", "subject_name", "subject_id", "birth_date",
    "birth_time", "birth_time_precision", "birth_place", "latitude", "longitude",
    "timezone", "birth_data_source", "birth_data_quality", "domain", "event_type",
    "event_start", "event_end", "event_time_precision", "event_description",
    "event_direction", "event_source", "event_verification_quality", "source_type",
    "source_title", "source_author", "source_publication", "source_page",
    "source_passage_reference", "original_case_source", "independent_verification",
    "prediction_cutoff", "knowledge_cutoff", "outcome_cutoff", "outcome_known_at_entry",
    "notes",
)

ALIASES = {
    "case id": "case_external_id", "case_id": "case_external_id", "case class": "case_class",
    "dob": "birth_date", "birthdate": "birth_date", "birth time": "birth_time",
    "place": "birth_place", "lat": "latitude", "lon": "longitude", "tz": "timezone",
    "event date": "event_start", "event date/window": "event_start", "event end": "event_end",
    "event verification": "event_verification_quality", "source": "source_title",
    "author": "source_author", "publisher": "source_publication", "page": "source_page",
    "passage": "source_passage_reference", "prediction cutoff": "prediction_cutoff",
    "knowledge cutoff": "knowledge_cutoff", "outcome cutoff": "outcome_cutoff",
}


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _fingerprint(value: Any) -> str:
    return hashlib.sha256(_json(value).encode("utf-8")).hexdigest()


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _parse_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def canonical_field(value: str) -> str:
    normalized = " ".join(value.strip().lower().replace("-", " ").replace("_", " ").split())
    return ALIASES.get(normalized, normalized.replace(" ", "_"))


def auto_mapping(headers: Iterable[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for header in headers:
        field = canonical_field(header)
        if field in TEMPLATE_FIELDS:
            mapping[header] = field
    return mapping


def _event_types() -> set[str]:
    return {item for values in EVENT_TYPES.values() for item in values} | {"NO_EVENT", "OTHER"}


def _candidate_case(row: dict[str, Any], *, import_id: str, row_number: int) -> CaseRecord:
    outcome = {
        "event_type": _clean(row.get("event_type")),
        "event_start": _clean(row.get("event_start")),
        "event_end": _clean(row.get("event_end")),
        "event_direction": _clean(row.get("event_direction")),
        "description": _clean(row.get("event_description")),
    }
    payload = {
        "case_id": _clean(row.get("case_external_id")) or None,
        "subject_id": _clean(row.get("subject_id")) or _clean(row.get("subject_name")) or f"IMPORT-{import_id}-{row_number}",
        "subject_label": _clean(row.get("subject_name")),
        "source_id": _clean(row.get("original_case_source")) or _clean(row.get("source_title")) or f"IMPORT-{import_id}",
        "source_type": _clean(row.get("source_type")),
        "source_title": _clean(row.get("source_title")), "author": _clean(row.get("source_author")),
        "publication": _clean(row.get("source_publication")), "source_page": _clean(row.get("source_page")),
        "original_case_source": _clean(row.get("original_case_source")),
        "independent_verification": _clean(row.get("independent_verification")),
        "passage_reference": _clean(row.get("source_passage_reference")) or "REFERENCE_NOT_VERIFIED",
        "domain": _clean(row.get("domain")).upper() or "GENERAL_TIMING",
        "case_class": _clean(row.get("case_class")).upper() or "UNVERIFIED",
        "chart_input": {key: _clean(row.get(key)) for key in ("birth_date", "birth_time", "birth_time_precision", "birth_place", "latitude", "longitude", "timezone", "birth_data_source", "birth_data_quality") if _clean(row.get(key))},
        "prediction_cutoff": _clean(row.get("prediction_cutoff")) or None,
        "knowledge_cutoff": _clean(row.get("knowledge_cutoff")) or None,
        "outcome_cutoff": _clean(row.get("outcome_cutoff")) or None,
        "outcome": outcome if outcome["event_type"] else None,
        "outcome_source": _clean(row.get("event_source")) or "UNVERIFIED",
        "verification_quality": _clean(row.get("event_verification_quality")).upper() or "UNVERIFIED",
        "birth_data_provenance": _clean(row.get("birth_data_source")).upper() or "UNVERIFIED",
        "event_provenance": _clean(row.get("event_verification_quality")).upper() or "UNVERIFIED",
        "case_family": _clean(row.get("case_family_id")) or None,
        "independent_source_family": _clean(row.get("independent_source_family")) or None,
        "notes": _clean(row.get("notes")),
    }
    return normalize_case(payload)


def _duplicate_state(case: CaseRecord, existing: Iterable[CaseRecord]) -> str:
    for item in existing:
        same_identity = case.subject_id == item.subject_id and case.chart_input.get("birth_date") and case.chart_input.get("birth_date") == item.chart_input.get("birth_date")
        same_event = (case.outcome or {}).get("event_type") == (item.outcome or {}).get("event_type") and (case.outcome or {}).get("event_start") == (item.outcome or {}).get("event_start")
        if case.case_family and case.case_family == item.case_family:
            return "SAME_CASE_FAMILY"
        if same_identity and same_event and case.source_id == item.source_id:
            return "EXACT_DUPLICATE"
        if same_identity and same_event:
            return "LIKELY_DUPLICATE"
    return "NO_DUPLICATE_FOUND"


def validate_case(case: CaseRecord, *, raw: dict[str, Any], duplicate_state: str = "NO_DUPLICATE_FOUND") -> dict[str, Any]:
    case.quality = assess_quality(case)
    errors: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []
    if not case.chart_input.get("birth_date"):
        errors.append({"code": "MISSING_BIRTH_DATE", "field": "birth_date", "message": "Birth date is required."})
    elif not _parse_date(case.chart_input["birth_date"]):
        errors.append({"code": "INVALID_DATE", "field": "birth_date", "message": "Use YYYY-MM-DD."})
    if not case.outcome or not case.outcome.get("event_type"):
        errors.append({"code": "MISSING_EVENT", "field": "event_type", "message": "At least one event type is required."})
    elif case.outcome["event_type"] not in _event_types():
        errors.append({"code": "UNMAPPED_EVENT", "field": "event_type", "message": "Event type is not in the governed taxonomy."})
    event_date = _parse_date(_clean((case.outcome or {}).get("event_start")))
    if (case.outcome or {}).get("event_start") and not event_date:
        errors.append({"code": "INVALID_DATE", "field": "event_start", "message": "Event date must use YYYY-MM-DD."})
    if case.case_class not in CASE_CLASSES:
        errors.append({"code": "INVALID_CASE_CLASS", "field": "case_class", "message": "Case class is not supported."})
    if not (_clean(raw.get("source_title")) or _clean(raw.get("source_type")) or _clean(raw.get("original_case_source"))):
        errors.append({"code": "MISSING_PROVENANCE", "field": "source_title", "message": "Source provenance is required."})
    if case.passage_reference == "REFERENCE_NOT_VERIFIED":
        warnings.append({"code": "REFERENCE_NOT_VERIFIED", "field": "source_passage_reference", "message": "Exact passage provenance is not verified."})
    if not case.chart_input.get("birth_time") or case.chart_input.get("birth_time_precision") not in {"EXACT", "EXACT_TIME"}:
        warnings.append({"code": "BIRTH_TIME_NOT_EXACT", "field": "birth_time_precision", "message": "Prediction precision may be limited."})
    historical = case.case_class.startswith("HISTORICAL") or case.case_class in {"WORKED_ASTROLOGY_CASE", "PRACTITIONER_CASE"}
    leakage = False
    if historical and not (case.prediction_cutoff and case.knowledge_cutoff and case.outcome_cutoff):
        warnings.append({"code": "CUTOFFS_REQUIRED", "field": "prediction_cutoff", "message": "Historical cases without complete cutoffs remain research-only."})
        leakage = True
    if case.prediction_cutoff and event_date and event_date <= (_parse_date(case.prediction_cutoff) or date.min):
        warnings.append({"code": "LEAKAGE_RISK", "field": "prediction_cutoff", "message": "The event is not after the prediction cutoff."})
        leakage = True
    if duplicate_state in {"EXACT_DUPLICATE", "LIKELY_DUPLICATE", "SAME_CASE_FAMILY"}:
        errors.append({"code": "DUPLICATE_CASE", "field": "case_family_id", "message": duplicate_state.replace("_", " ").title()})
    if duplicate_state != "NO_DUPLICATE_FOUND":
        eligibility = "DUPLICATE"
    elif errors:
        eligibility = "INCOMPLETE" if any(item["code"] in {"MISSING_BIRTH_DATE", "MISSING_EVENT", "MISSING_PROVENANCE"} for item in errors) else "REJECTED"
    elif leakage:
        eligibility = "LEAKAGE_INVALID"
    elif case.case_class in {"HISTORICAL_VERIFIED", "PROSPECTIVE_VERIFIED"} and case.quality in {"HIGH", "MODERATE"}:
        eligibility = "ELIGIBLE"
    elif case.case_class in {"HISTORICAL_DOCUMENTED", "WORKED_ASTROLOGY_CASE", "PRACTITIONER_CASE", "HISTORICAL_USER_REPORTED"}:
        eligibility = "RESEARCH_ONLY"
    else:
        eligibility = "UNVERIFIED"
    return {"status": "ERROR" if errors else "WARNING" if warnings else "VALID", "errors": errors, "warnings": warnings, "eligibility": eligibility, "duplicate_state": duplicate_state, "quality": case.quality}


class CaseIntakeService:
    """Intake orchestration; raw cases remain in pred_cases on the PRED store."""

    def __init__(self, db_path: str | Path | None = None) -> None:
        self.db_path = Path(db_path or cfg.VEDA_RESEARCH_PLATFORM_DB)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.registry = CaseRegistry(self.db_path)
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        con = sqlite3.connect(str(self.db_path), timeout=30)
        con.row_factory = sqlite3.Row
        con.execute("PRAGMA foreign_keys=ON")
        return con

    def _init_db(self) -> None:
        with self._connect() as con:
            con.executescript("""
            CREATE TABLE IF NOT EXISTS empirical_imports (
                import_id TEXT PRIMARY KEY, filename TEXT NOT NULL, file_type TEXT NOT NULL,
                content_fingerprint TEXT NOT NULL UNIQUE, uploaded_at TEXT NOT NULL,
                uploaded_by TEXT NOT NULL, row_count INTEGER NOT NULL, mapping TEXT NOT NULL,
                status TEXT NOT NULL, summary TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS empirical_import_rows (
                import_id TEXT NOT NULL, row_number INTEGER NOT NULL, status TEXT NOT NULL,
                severity TEXT NOT NULL, payload TEXT NOT NULL, validation TEXT NOT NULL,
                case_id TEXT, PRIMARY KEY(import_id, row_number)
            );
            CREATE TABLE IF NOT EXISTS empirical_case_audit (
                audit_id TEXT PRIMARY KEY, case_id TEXT, import_id TEXT, action TEXT NOT NULL,
                actor TEXT NOT NULL, created_at TEXT NOT NULL, payload TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_empirical_imports_time ON empirical_imports(uploaded_at);
            CREATE INDEX IF NOT EXISTS idx_empirical_case_audit_case ON empirical_case_audit(case_id, created_at);
            """)

    def _existing_cases(self) -> list[CaseRecord]:
        with self._connect() as con:
            rows = con.execute("SELECT payload FROM pred_cases").fetchall()
        return [normalize_case(json.loads(row["payload"])) for row in rows]

    def validate_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        case = _candidate_case(payload, import_id="SINGLE", row_number=1)
        duplicate = _duplicate_state(case, self._existing_cases())
        result = validate_case(case, raw=payload, duplicate_state=duplicate)
        result["case"] = case.to_dict()
        return result

    def create_case(self, payload: dict[str, Any], *, actor: str) -> dict[str, Any]:
        result = self.validate_payload(payload)
        if result["status"] == "ERROR":
            raise ValueError(result)
        case = _candidate_case(payload, import_id="SINGLE", row_number=1)
        case.quality = result["quality"]
        case.leakage_status = "VALID" if result["eligibility"] in {"ELIGIBLE", "ELIGIBLE_WITH_CONDITIONS", "RESEARCH_ONLY", "UNVERIFIED"} else result["eligibility"]
        added, status = self.registry.add(case)
        if status != "ADDED":
            result["status"] = "DUPLICATE"
            result["eligibility"] = "DUPLICATE"
        self._audit(case_id=added.case_id, action="CASE_CREATED" if status == "ADDED" else "CASE_DUPLICATE", actor=actor, payload={"validation": result, "status": status})
        return {"case": added.to_dict(), "status": status, "validation": result}

    def _audit(self, *, case_id: str | None, action: str, actor: str, payload: dict[str, Any], import_id: str | None = None) -> None:
        audit_payload = {"case_id": case_id, "import_id": import_id, "action": action, "payload": payload}
        audit_id = f"AUDIT-{_fingerprint(audit_payload)[:20]}"
        with self._connect() as con:
            con.execute("INSERT OR IGNORE INTO empirical_case_audit VALUES (?,?,?,?,?,?,?)", (audit_id, case_id, import_id, action, actor, _now(), _json(payload)))

    def _parse_rows(self, content: bytes, filename: str, sheet: str | None = None) -> tuple[list[dict[str, Any]], list[str], str]:
        suffix = Path(filename).suffix.lower()
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            headers = list(reader.fieldnames or [])
            return [dict(row) for row in reader], headers, "CSV"
        if suffix == ".xlsx":
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            worksheet = workbook[sheet] if sheet and sheet in workbook.sheetnames else workbook[workbook.sheetnames[0]]
            values = list(worksheet.values)
            headers = [str(value or "") for value in (values[0] if values else [])]
            rows = [dict(zip(headers, ("" if value is None else value for value in row))) for row in values[1:]]
            return rows, headers, "XLSX"
        raise ValueError("Only CSV and XLSX files are supported.")

    def preview_import(self, content: bytes, filename: str, *, actor: str, mapping: dict[str, str] | None = None, sheet: str | None = None) -> dict[str, Any]:
        if len(content) > MAX_UPLOAD_BYTES:
            raise ValueError("Upload exceeds the 10 MB limit.")
        fingerprint = hashlib.sha256(content).hexdigest()
        rows, headers, file_type = self._parse_rows(content, filename, sheet)
        mapping = mapping or auto_mapping(headers)
        normalized_rows = [{mapping.get(header, canonical_field(header)): value for header, value in row.items() if mapping.get(header, canonical_field(header)) in TEMPLATE_FIELDS} for row in rows]
        import_id = "IMP-" + fingerprint[:16]
        existing = self._existing_cases()
        results = []
        for index, row in enumerate(normalized_rows, start=1):
            case = _candidate_case(row, import_id=import_id, row_number=index)
            validation = validate_case(case, raw=row, duplicate_state=_duplicate_state(case, existing))
            results.append({"row_number": index, "status": validation["status"], "severity": validation["status"], "validation": validation, "payload": row, "case_id": case.case_id})
        summary = self._summary(results)
        with self._connect() as con:
            con.execute("INSERT OR REPLACE INTO empirical_imports VALUES (?,?,?,?,?,?,?,?,?,?)", (import_id, filename, file_type, fingerprint, _now(), actor, len(results), _json(mapping), "READY_FOR_REVIEW", _json(summary)))
            con.execute("DELETE FROM empirical_import_rows WHERE import_id=?", (import_id,))
            for item in results:
                con.execute("INSERT INTO empirical_import_rows VALUES (?,?,?,?,?,?,?)", (import_id, item["row_number"], item["status"], item["severity"], _json(item["payload"]), _json(item["validation"]), item["case_id"]))
        self._audit(case_id=None, import_id=import_id, action="IMPORT_PREVIEWED", actor=actor, payload={"filename": filename, "summary": summary, "mapping": mapping})
        return {"import_id": import_id, "filename": filename, "file_type": file_type, "headers": headers, "mapping": mapping, "summary": summary, "rows": results, "sheet": sheet}

    def _summary(self, results: list[dict[str, Any]]) -> dict[str, int]:
        return {
            "rows": len(results), "valid": sum(item["status"] == "VALID" for item in results),
            "warnings": sum(bool(item["validation"]["warnings"]) for item in results),
            "errors": sum(bool(item["validation"]["errors"]) for item in results),
            "duplicates": sum(item["validation"]["duplicate_state"] != "NO_DUPLICATE_FOUND" for item in results),
            "eligible": sum(item["validation"]["eligibility"] == "ELIGIBLE" for item in results),
            "research_only": sum(item["validation"]["eligibility"] == "RESEARCH_ONLY" for item in results),
            "rejected": sum(item["validation"]["eligibility"] in {"REJECTED", "INCOMPLETE", "LEAKAGE_INVALID"} for item in results),
        }

    def ingest_import(self, import_id: str, *, actor: str, rows: list[int] | None = None) -> dict[str, Any]:
        with self._connect() as con:
            import_row = con.execute("SELECT * FROM empirical_imports WHERE import_id=?", (import_id,)).fetchone()
            if not import_row:
                raise KeyError(import_id)
            stored_rows = con.execute("SELECT * FROM empirical_import_rows WHERE import_id=? ORDER BY row_number", (import_id,)).fetchall()
        selected = set(rows) if rows else None
        accepted = rejected = duplicates = 0
        for item in stored_rows:
            if selected is not None and item["row_number"] not in selected:
                continue
            validation = json.loads(item["validation"])
            if validation["status"] == "ERROR" or validation["eligibility"] in {"DUPLICATE", "INCOMPLETE", "REJECTED", "LEAKAGE_INVALID"}:
                rejected += 1
                continue
            if any(existing.case_id == item["case_id"] for existing in self._existing_cases()):
                duplicates += 1
                continue
            result = self.create_case(json.loads(item["payload"]), actor=actor)
            if result["status"] == "ADDED":
                accepted += 1
            else:
                duplicates += 1
        status = "INGESTED" if rejected == 0 else "PARTIALLY_INGESTED" if accepted else "REJECTED"
        summary = {"accepted": accepted, "rejected": rejected, "duplicates": duplicates, "status": status}
        with self._connect() as con:
            con.execute("UPDATE empirical_imports SET status=?, summary=? WHERE import_id=?", (status, _json(summary), import_id))
        self._audit(case_id=None, import_id=import_id, action="IMPORT_INGESTED", actor=actor, payload=summary)
        return {"import_id": import_id, **summary}

    def list_imports(self, limit: int = 50) -> list[dict[str, Any]]:
        with self._connect() as con:
            rows = con.execute("SELECT * FROM empirical_imports ORDER BY uploaded_at DESC LIMIT ?", (limit,)).fetchall()
        return [{**dict(row), "mapping": json.loads(row["mapping"]), "summary": json.loads(row["summary"])} for row in rows]

    def import_detail(self, import_id: str) -> dict[str, Any]:
        with self._connect() as con:
            item = con.execute("SELECT * FROM empirical_imports WHERE import_id=?", (import_id,)).fetchone()
            rows = con.execute("SELECT * FROM empirical_import_rows WHERE import_id=? ORDER BY row_number", (import_id,)).fetchall()
        if not item:
            raise KeyError(import_id)
        return {**dict(item), "mapping": json.loads(item["mapping"]), "summary": json.loads(item["summary"]), "rows": [{**dict(row), "payload": json.loads(row["payload"]), "validation": json.loads(row["validation"])} for row in rows]}

    def list_cases(self, limit: int = 100) -> list[dict[str, Any]]:
        with self._connect() as con:
            rows = con.execute("SELECT case_id, case_class, quality, leakage_status, payload FROM pred_cases ORDER BY rowid DESC LIMIT ?", (limit,)).fetchall()
        return [{**json.loads(row["payload"]), "case_id": row["case_id"], "quality": row["quality"], "leakage_status": row["leakage_status"]} for row in rows]

    def case_detail(self, case_id: str) -> dict[str, Any]:
        with self._connect() as con:
            row = con.execute("SELECT payload FROM pred_cases WHERE case_id=?", (case_id,)).fetchone()
            audits = con.execute("SELECT * FROM empirical_case_audit WHERE case_id=? ORDER BY created_at DESC", (case_id,)).fetchall()
        if not row:
            raise KeyError(case_id)
        case = json.loads(row["payload"])
        case["audit_history"] = [{**dict(item), "payload": json.loads(item["payload"])} for item in audits]
        return case

    def counts(self) -> dict[str, int]:
        with self._connect() as con:
            total = con.execute("SELECT COUNT(*) FROM pred_cases").fetchone()[0]
            eligible = con.execute("SELECT COUNT(*) FROM pred_cases WHERE case_class IN ('HISTORICAL_VERIFIED','PROSPECTIVE_VERIFIED') AND quality IN ('HIGH','MODERATE') AND leakage_status='VALID'").fetchone()[0]
        return {"cases": int(total), "eligible": int(eligible)}


def template_csv() -> str:
    return ",".join(TEMPLATE_FIELDS) + "\n"


__all__ = ["CaseIntakeService", "TEMPLATE_FIELDS", "TEMPLATE_VERSION", "auto_mapping", "validate_case", "template_csv"]
